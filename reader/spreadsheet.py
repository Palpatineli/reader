from typing import Tuple, Union

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

Point = Tuple[int, int]


def detect_block(ws: Worksheet, start_x: Union[str, int],
                 start_y: int) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    """Find out the extend of one connected rectangular data block
    Args:
        ws: the worksheet to look on
        start_x: x coordinate of one cell inside the block
        start_y: y coordinate of one cell inside the block
    Returns:
        (start_point, end_point)
    """
    start_x = (column_index_from_string(start_x) if isinstance(start_x, str) else start_x) - 1
    if not ws[start_y][start_x].value:
        raise ValueError('{0}, {1} is not inside a data block!')
    end_x, end_y = start_x + 1, start_y + 1
    while True:
        try:
            if start_y > 1 and any(x.value for x in ws[start_y - 1][start_x: end_x]):
                start_y -= 1
            elif any(x.value for x in ws[end_y][start_x: end_x]):
                end_y += 1
            elif start_x > 0 and any(line[start_x - 1].value for line in ws[start_y: end_y]):
                start_x -= 1
            elif any((end_x < len(line) and line[end_x].value) for line in ws[start_y: end_y]):
                end_x += 1
            else:
                break
        except IndexError as e:
            print(start_x, start_y, end_x, end_y)
            raise e
    return (start_x, start_y), (end_x, end_y)


def extract_data(file_path: str, start_x: Union[str, int], start_y: int) -> Tuple[list, np.ndarray]:
    ws = load_workbook(file_path).active
    (x_start, y_start), (x_end, y_end) = detect_block(ws, start_x, start_y)
    header = [x.value for x in ws[y_start][x_start: x_end]]
    data = np.array([[x.value if x.value else np.nan for x in line[x_start: x_end]]
                     for line in ws[y_start + 1: y_end - 1]])
    return header, data
